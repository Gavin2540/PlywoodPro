import datetime
import os
import sqlite3
from database.db import get_db

# openpyxl for Excel exports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# reportlab for PDF exports
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.units import cm

import models.sale as sale_model
import models.expense as expense_model
import models.ledger as ledger_model
import utils.helpers as helpers

def export_daily_excel(date_str, filepath):
    """Generates a highly-stylized, professional Excel spreadsheet workbook
    containing Daily Financial Summaries, Sales Margin Logs, Expenses, and Stock Reconciliation Ledger.
    """
    # Fetch Data
    sales = sale_model.get_sales_by_date(date_str)
    expenses = expense_model.get_expenses_by_date(date_str)
    ledger = ledger_model.get_ledger_for_date(date_str)

    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # STYLE DEFINITIONS
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    kpi_title_font = Font(name=font_family, size=9, bold=True, color="808080")
    kpi_value_font = Font(name=font_family, size=14, bold=True, color="000000")
    bold_font = Font(name=font_family, size=10, bold=True, color="000000")
    regular_font = Font(name=font_family, size=10, color="000000")
    italic_font = Font(name=font_family, size=10, italic=True, color="555555")

    # Fills
    brand_fill = PatternFill(start_color="BA7517", end_color="BA7517", fill_type="solid") # Amber
    header_fill = PatternFill(start_color="1F1E1B", end_color="1F1E1B", fill_type="solid") # Dark Charcoal
    zebra_fill = PatternFill(start_color="F9F9F8", end_color="F9F9F8", fill_type="solid") # Light grayish alternate
    kpi_fill = PatternFill(start_color="F4F2EE", end_color="F4F2EE", fill_type="solid") # Soft card fill
    profit_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid") # Soft green
    loss_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid") # Soft red

    # Borders
    thin_border_side = Side(border_style="thin", color="CCCCCC")
    double_border_side = Side(border_style="double", color="333333")
    thick_border_side = Side(border_style="medium", color="BA7517")
    
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    summary_border = Border(top=thin_border_side, bottom=double_border_side)
    kpi_border = Border(left=thin_border_side, right=thin_border_side, top=thick_border_side, bottom=thin_border_side)

    # ----------------------------------------------------
    # SHEET 1: Financial & Sales Summary
    # ----------------------------------------------------
    ws_sales = wb.active
    ws_sales.title = "Sales & Margins"
    ws_sales.views.sheetView[0].showGridLines = True

    # Title block
    ws_sales.merge_cells("A1:M1")
    title_cell = ws_sales["A1"]
    title_cell.value = f"DAILY SALES & PROFITS MARGIN BREAKDOWN — {helpers.format_date(date_str)}"
    title_cell.font = title_font
    title_cell.fill = brand_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_sales.row_dimensions[1].height = 40

    # Stat Cards / KPI Blocks
    total_qty = sum(s['qty'] for s in sales)
    total_rev = sum(s['total_revenue'] for s in sales)
    total_gp = sum(s['margin_profit'] for s in sales)
    total_exp = sum(e['amount'] for e in expenses)
    net_profit = total_gp - total_exp

    kpi_data = [
        ("Volume Sold", f"{total_qty:.1f} Units", "B3:C4"),
        ("Gross Sales Revenue", total_rev, "E3:F4"),
        ("Gross Margin Profit", total_gp, "H3:I4"),
        ("Daily Expenditures", total_exp, "K3:L4")
    ]

    for label, val, cells_range in kpi_data:
        start_cell_ref, end_cell_ref = cells_range.split(":")
        start_col = start_cell_ref[0]
        end_col = end_cell_ref[0]
        
        # Label in top row (row 3)
        ws_sales[f"{start_col}3"].value = label.upper()
        ws_sales[f"{start_col}3"].font = kpi_title_font
        ws_sales[f"{start_col}3"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sales.merge_cells(f"{start_col}3:{end_col}3")
        
        # Value in bottom row (row 4)
        val_cell = ws_sales[f"{start_col}4"]
        val_cell.value = val
        val_cell.font = kpi_value_font
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_sales.merge_cells(f"{start_col}4:{end_col}4")
        
        if isinstance(val, (int, float)):
            val_cell.number_format = '"₹"#,##,##0.00'
        
        # Apply border & fills
        start_c, start_r = openpyxl.utils.coordinate_to_tuple(start_cell_ref)
        end_c, end_r = openpyxl.utils.coordinate_to_tuple(end_cell_ref)
        for r in range(start_r, end_r + 1):
            for c in range(start_c, end_c + 1):
                cell = ws_sales.cell(row=r, column=c)
                cell.fill = kpi_fill
                cell.border = kpi_border
                
    ws_sales.row_dimensions[3].height = 18
    ws_sales.row_dimensions[4].height = 25
    
    # Net Profit Card Banner
    ws_sales.merge_cells("B5:L5")
    np_cell = ws_sales["B5"]
    np_cell.value = f"TOTAL DAILY NET BUSINESS PROFIT:   " + helpers.format_currency(net_profit)
    np_cell.font = Font(name=font_family, size=11, bold=True, color="000000")
    np_cell.alignment = Alignment(horizontal="center", vertical="center")
    np_cell.fill = profit_fill if net_profit >= 0 else loss_fill
    ws_sales.row_dimensions[5].height = 25
    for c in range(2, 13):
        ws_sales.cell(row=5, column=c).border = cell_border
        ws_sales.cell(row=5, column=c).fill = profit_fill if net_profit >= 0 else loss_fill

    # Sales Table Header
    sales_headers = [
        "Product Name", "Brand", "Type", "Thickness", "Size", "Qty Sold", "Unit", 
        "Selling Rate", "Cost Rate", "Total Revenue", "Total Cost Price", "Gross Margin Profit", "Customer & Notes"
    ]
    
    start_row_sales = 8
    ws_sales.row_dimensions[start_row_sales].height = 25
    
    for idx, name in enumerate(sales_headers, start=1):
        cell = ws_sales.cell(row=start_row_sales, column=idx)
        cell.value = name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if idx in [1,2,13] else "right", vertical="center")
        cell.border = cell_border

    # Data Rows
    current_row = start_row_sales + 1
    for idx, s in enumerate(sales):
        ws_sales.row_dimensions[current_row].height = 20
        fill_row = zebra_fill if idx % 2 == 1 else PatternFill(fill_type=None)
        
        # Product specifications
        ws_sales.cell(row=current_row, column=1, value=s['product_name']).font = bold_font
        ws_sales.cell(row=current_row, column=2, value=s['product_brand']).font = regular_font
        ws_sales.cell(row=current_row, column=3, value=s['product_type']).font = regular_font
        ws_sales.cell(row=current_row, column=4, value=s['product_thickness']).font = regular_font
        ws_sales.cell(row=current_row, column=5, value=s['product_size']).font = regular_font
        
        # Quantity & Unit
        ws_sales.cell(row=current_row, column=6, value=s['qty']).font = bold_font
        ws_sales.cell(row=current_row, column=6).number_format = '#,##0.0'
        ws_sales.cell(row=current_row, column=7, value=s['product_unit']).font = regular_font
        
        # Prices
        ws_sales.cell(row=current_row, column=8, value=s['unit_price']).font = regular_font
        ws_sales.cell(row=current_row, column=8).number_format = '"₹"#,##,##0.00'
        ws_sales.cell(row=current_row, column=9, value=s['purchase_price_at_time']).font = regular_font
        ws_sales.cell(row=current_row, column=9).number_format = '"₹"#,##,##0.00'
        
        # Computations (via Excel formulas for clean values!)
        # Total Revenue = Qty * Selling Rate
        ws_sales.cell(row=current_row, column=10, value=f"=F{current_row}*H{current_row}").font = bold_font
        ws_sales.cell(row=current_row, column=10).number_format = '"₹"#,##,##0.00'
        
        # Total Cost = Qty * Cost Rate
        ws_sales.cell(row=current_row, column=11, value=f"=F{current_row}*I{current_row}").font = regular_font
        ws_sales.cell(row=current_row, column=11).number_format = '"₹"#,##,##0.00'
        
        # Profit Margin = Total Revenue - Total Cost
        ws_sales.cell(row=current_row, column=12, value=f"=J{current_row}-K{current_row}").font = bold_font
        ws_sales.cell(row=current_row, column=12).number_format = '"₹"#,##,##0.00'
        
        # Customer Note
        note_str = s['customer_name'] or ""
        if s['notes']:
            note_str += f" ({s['notes']})"
        ws_sales.cell(row=current_row, column=13, value=note_str).font = italic_font
        
        # Alignments & Borders
        for c in range(1, 14):
            cell = ws_sales.cell(row=current_row, column=c)
            cell.border = cell_border
            if fill_row.fill_type:
                cell.fill = fill_row
            if c not in [1, 2, 13]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    # Total row at bottom
    if sales:
        ws_sales.row_dimensions[current_row].height = 24
        # Total text
        ws_sales.cell(row=current_row, column=1, value="GRAND TOTALS").font = Font(name=font_family, size=10, bold=True, color="BA7517")
        # Formula sums
        ws_sales.cell(row=current_row, column=6, value=f"=SUM(F9:F{current_row-1})").font = bold_font
        ws_sales.cell(row=current_row, column=6).number_format = '#,##0.0'
        
        ws_sales.cell(row=current_row, column=10, value=f"=SUM(J9:J{current_row-1})").font = bold_font
        ws_sales.cell(row=current_row, column=10).number_format = '"₹"#,##,##0.00'
        
        ws_sales.cell(row=current_row, column=11, value=f"=SUM(K9:K{current_row-1})").font = bold_font
        ws_sales.cell(row=current_row, column=11).number_format = '"₹"#,##,##0.00'
        
        ws_sales.cell(row=current_row, column=12, value=f"=SUM(L9:L{current_row-1})").font = bold_font
        ws_sales.cell(row=current_row, column=12).number_format = '"₹"#,##,##0.00'
        
        for c in range(1, 14):
            cell = ws_sales.cell(row=current_row, column=c)
            cell.border = summary_border
            cell.fill = kpi_fill
            if c not in [1, 2, 13]:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # ----------------------------------------------------
    # SHEET 2: Operating Expenditures
    # ----------------------------------------------------
    ws_exp = wb.create_sheet(title="Operating Expenses")
    ws_exp.views.sheetView[0].showGridLines = True

    # Title Block
    ws_exp.merge_cells("A1:C1")
    title_exp = ws_exp["A1"]
    title_exp.value = f"DAILY BUSINESS OPERATING EXPENDITURES — {helpers.format_date(date_str)}"
    title_exp.font = title_font
    title_exp.fill = brand_fill
    title_exp.alignment = Alignment(horizontal="center", vertical="center")
    ws_exp.row_dimensions[1].height = 40

    # Table Headers
    exp_headers = ["Expenditure Category", "Amount", "Note / Detailed Explanation"]
    ws_exp.row_dimensions[3].height = 25
    for idx, name in enumerate(exp_headers, start=1):
        cell = ws_exp.cell(row=3, column=idx)
        cell.value = name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if idx in [1,3] else "right", vertical="center")
        cell.border = cell_border

    # Data Rows
    current_row_exp = 4
    for idx, e in enumerate(expenses):
        ws_exp.row_dimensions[current_row_exp].height = 20
        fill_row = zebra_fill if idx % 2 == 1 else PatternFill(fill_type=None)
        
        ws_exp.cell(row=current_row_exp, column=1, value=e['category']).font = bold_font
        ws_exp.cell(row=current_row_exp, column=2, value=e['amount']).font = Font(name=font_family, size=10, bold=True, color="C0392B") # red text
        ws_exp.cell(row=current_row_exp, column=2).number_format = '"₹"#,##,##0.00'
        ws_exp.cell(row=current_row_exp, column=3, value=e['note'] or "-").font = regular_font
        
        for c in range(1, 4):
            cell = ws_exp.cell(row=current_row_exp, column=c)
            cell.border = cell_border
            if fill_row.fill_type:
                cell.fill = fill_row
            if c == 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row_exp += 1

    # Total Row
    if expenses:
        ws_exp.row_dimensions[current_row_exp].height = 24
        ws_exp.cell(row=current_row_exp, column=1, value="TOTAL OPERATING EXPENDITURES").font = Font(name=font_family, size=10, bold=True, color="C0392B")
        ws_exp.cell(row=current_row_exp, column=2, value=f"=SUM(B4:B{current_row_exp-1})").font = Font(name=font_family, size=11, bold=True, color="C0392B")
        ws_exp.cell(row=current_row_exp, column=2).number_format = '"₹"#,##,##0.00'
        
        for c in range(1, 4):
            cell = ws_exp.cell(row=current_row_exp, column=c)
            cell.border = summary_border
            cell.fill = kpi_fill
            if c == 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # ----------------------------------------------------
    # SHEET 3: Stock Ledger Reconciliation
    # ----------------------------------------------------
    ws_led = wb.create_sheet(title="Inventory Reconciliation")
    ws_led.views.sheetView[0].showGridLines = True

    # Title Block
    ws_led.merge_cells("A1:I1")
    title_led = ws_led["A1"]
    title_led.value = f"INVENTORY RECONCILIATION & STOCK LEDGER — {helpers.format_date(date_str)}"
    title_led.font = title_font
    title_led.fill = brand_fill
    title_led.alignment = Alignment(horizontal="center", vertical="center")
    ws_led.row_dimensions[1].height = 40

    # Headers
    led_headers = [
        "Product Specifications", "Brand", "Thickness", "Dimensions Size", 
        "Opening Stock", "Purchases (+)", "Sales (-)", "Closing Stock", "Reconciliation Status"
    ]
    ws_led.row_dimensions[3].height = 25
    for idx, name in enumerate(led_headers, start=1):
        cell = ws_led.cell(row=3, column=idx)
        cell.value = name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left" if idx in [1,2,9] else "right", vertical="center")
        cell.border = cell_border

    # Data Rows
    current_row_led = 4
    for idx, row in enumerate(ledger):
        ws_led.row_dimensions[current_row_led].height = 20
        fill_row = zebra_fill if idx % 2 == 1 else PatternFill(fill_type=None)
        
        # Product Specs
        ws_led.cell(row=current_row_led, column=1, value=row['product_name']).font = bold_font
        ws_led.cell(row=current_row_led, column=2, value=row['product_brand']).font = regular_font
        ws_led.cell(row=current_row_led, column=3, value=row['product_thickness']).font = regular_font
        ws_led.cell(row=current_row_led, column=4, value=row['product_size']).font = regular_font
        
        # Quantities
        ws_led.cell(row=current_row_led, column=5, value=row['opening_stock']).font = regular_font
        ws_led.cell(row=current_row_led, column=5).number_format = f'#,##0.0" {row["product_unit"]}"'
        
        ws_led.cell(row=current_row_led, column=6, value=row['purchases_qty']).font = regular_font
        ws_led.cell(row=current_row_led, column=6).number_format = f'#,##0.0" {row["product_unit"]}"'
        
        ws_led.cell(row=current_row_led, column=7, value=row['sales_qty']).font = regular_font
        ws_led.cell(row=current_row_led, column=7).number_format = f'#,##0.0" {row["product_unit"]}"'
        
        # Closing Stock (Formula-based or overridden value)
        if row['manual_override'] == 1:
            ws_led.cell(row=current_row_led, column=8, value=row['closing_stock']).font = Font(name=font_family, size=10, bold=True, color="3498DB")
        else:
            ws_led.cell(row=current_row_led, column=8, value=f"=E{current_row_led}+F{current_row_led}-G{current_row_led}").font = bold_font
        ws_led.cell(row=current_row_led, column=8).number_format = f'#,##0.0" {row["product_unit"]}"'
        
        # Status Label
        if row['is_confirmed'] == 1:
            status_desc = "LOCKED"
            status_text_color = "27AE60" # green
        elif row['manual_override'] == 1:
            status_desc = f"ADJUSTED"
            if row['override_note']:
                status_desc += f" ({row['override_note']})"
            status_text_color = "2980B9" # blue
        else:
            status_desc = "UNLOCKED"
            status_text_color = "D35400" # orange
            
        status_cell = ws_led.cell(row=current_row_led, column=9, value=status_desc)
        status_cell.font = Font(name=font_family, size=9, bold=True, color=status_text_color)
        
        for c in range(1, 10):
            cell = ws_led.cell(row=current_row_led, column=c)
            cell.border = cell_border
            if fill_row.fill_type:
                cell.fill = fill_row
            if c not in [1, 2, 9]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row_led += 1

    # ----------------------------------------------------
    # COLUMN WIDTH AUTO-FITTER (ALL SHEETS)
    # ----------------------------------------------------
    for ws in [ws_sales, ws_exp, ws_led]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Skip checking values in row 1 & row 3-4 (KPIs) to prevent oversized columns
            for cell in col:
                if cell.row in [1, 3, 4, 5]:
                    continue
                val_str = str(cell.value or "")
                if cell.number_format and "₹" in cell.number_format:
                    val_str = "Rs." + val_str # rough length fallback
                max_len = max(max_len, len(val_str))
                
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Save Workbook
    wb.save(filepath)
    return True


def export_daily_pdf(date_str, filepath):
    """Generates a high-fidelity, styled business PDF summary report utilizing ReportLab.
    Ensures safe font representations (uses 'Rs.' in place of unicode '₹' to prevent system rendering errors).
    """
    # Fetch Data
    sales = sale_model.get_sales_by_date(date_str)
    expenses = expense_model.get_expenses_by_date(date_str)
    ledger = ledger_model.get_ledger_for_date(date_str)

    # Computations
    total_qty = sum(s['qty'] for s in sales)
    total_rev = sum(s['total_revenue'] for s in sales)
    total_gp = sum(s['margin_profit'] for s in sales)
    total_exp = sum(e['amount'] for e in expenses)
    net_profit = total_gp - total_exp

    # Set up Document
    # A4 Page dimensions: 595.27 x 841.89 points
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Palette
    c_brand = colors.HexColor("#BA7517")
    c_dark = colors.HexColor("#1F1E1B")
    c_light = colors.HexColor("#F9F9F8")
    c_border = colors.HexColor("#E5E2DB")
    c_accent_green = colors.HexColor("#27AE60")
    c_accent_red = colors.HexColor("#C0392B")

    # Typography Styles
    style_letterhead = ParagraphStyle(
        'Letterhead', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=18, textColor=colors.white,
        alignment=1, spaceAfter=2
    )
    
    style_subtitle = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontName='Helvetica-Oblique', fontSize=11, textColor=colors.HexColor("#DCDAD5"),
        alignment=1
    )

    style_h1 = ParagraphStyle(
        'H1', parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=12, textColor=c_dark,
        spaceBefore=14, spaceAfter=6, keepWithNext=True
    )

    style_cell_header = ParagraphStyle(
        'CellHeader', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8, textColor=colors.white,
        alignment=0
    )
    
    style_cell_header_right = ParagraphStyle(
        'CellHeaderRight', parent=style_cell_header,
        alignment=2
    )

    style_cell_normal = ParagraphStyle(
        'CellNormal', parent=styles['Normal'],
        fontName='Helvetica', fontSize=8, textColor=colors.HexColor("#2C2A29"),
        leading=10
    )
    
    style_cell_bold = ParagraphStyle(
        'CellBold', parent=style_cell_normal,
        fontName='Helvetica-Bold'
    )
    
    style_cell_right = ParagraphStyle(
        'CellRight', parent=style_cell_normal,
        alignment=2
    )

    style_cell_right_bold = ParagraphStyle(
        'CellRightBold', parent=style_cell_right,
        fontName='Helvetica-Bold'
    )

    style_kpi_label = ParagraphStyle(
        'KpiLabel', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor("#808080"),
        alignment=1
    )

    style_kpi_value = ParagraphStyle(
        'KpiValue', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=12, textColor=colors.black,
        alignment=1
    )

    story = []

    # ----------------------------------------------------
    # Header Banner (Letterhead)
    # ----------------------------------------------------
    header_data = [
        [Paragraph("PLYWOODPRO INVENTORY SYSTEM", style_letterhead)],
        [Paragraph(f"Daily Business Performance & Reconciliation Report — {helpers.format_date(date_str)}", style_subtitle)]
    ]
    header_table = Table(header_data, colWidths=[18*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), c_brand),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 0),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,1), (-1,-1), 10),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.4*cm))

    # ----------------------------------------------------
    # KPI Performance Summary Block
    # ----------------------------------------------------
    kpi_cells = [
        [
            Paragraph("VOLUME SOLD", style_kpi_label),
            Paragraph("GROSS SALES REVENUE", style_kpi_label),
            Paragraph("GROSS MARGIN PROFIT", style_kpi_label),
            Paragraph("OPERATING COSTS", style_kpi_label)
        ],
        [
            Paragraph(f"{total_qty:.1f} Units" if total_qty % 1 != 0 else f"{int(total_qty)} Units", style_kpi_value),
            Paragraph(f"Rs. {helpers.format_currency(total_rev)[1:]}", style_kpi_value),
            Paragraph(f"Rs. {helpers.format_currency(total_gp)[1:]}", style_kpi_value),
            Paragraph(f"Rs. {helpers.format_currency(total_exp)[1:]}", style_kpi_value)
        ]
    ]
    kpi_table = Table(kpi_cells, colWidths=[4.5*cm, 4.5*cm, 4.5*cm, 4.5*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor("#F2EFE9")),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 1, colors.white),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('BOTTOMPADDING', (0,0), (-1,0), 2),
        ('TOPPADDING', (0,1), (-1,1), 2),
        ('BOTTOMPADDING', (0,1), (-1,1), 6),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.2*cm))

    # Net Profit Banner
    np_color = c_accent_green if net_profit >= 0 else c_accent_red
    np_text = f"TOTAL DAILY NET BUSINESS PROFIT:   Rs. {helpers.format_currency(net_profit)[1:]}"
    if net_profit < 0:
         np_text = f"TOTAL DAILY NET BUSINESS LOSS:  -Rs. {helpers.format_currency(abs(net_profit))[1:]}"
         
    style_np_banner = ParagraphStyle(
        'NpBanner', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=10, textColor=colors.white,
        alignment=1
    )
    np_table = Table([[Paragraph(np_text, style_np_banner)]], colWidths=[18*cm])
    np_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), np_color),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(np_table)
    story.append(Spacer(1, 0.4*cm))

    # ----------------------------------------------------
    # SECTION 1: Itemized Sales Breakdown Table
    # ----------------------------------------------------
    story.append(Paragraph("1. ITEMIZED SALES & MARGIN BREAKDOWN", style_h1))
    
    sales_rows = [[
        Paragraph("Product Description & Size", style_cell_header),
        Paragraph("Qty", style_cell_header_right),
        Paragraph("Sale Rate", style_cell_header_right),
        Paragraph("Cost Rate", style_cell_header_right),
        Paragraph("Revenue", style_cell_header_right),
        Paragraph("Margin Profit", style_cell_header_right)
    ]]
    
    for idx, s in enumerate(sales):
        desc = f"{s['product_name']} ({s['product_brand']}) - {s['product_thickness']}, {s['product_size']}"
        qty_text = f"{s['qty']:.1f}" if s['qty'] % 1 != 0 else f"{int(s['qty'])}"
        
        row_cells = [
            Paragraph(desc, style_cell_normal),
            Paragraph(f"{qty_text} {s['product_unit']}", style_cell_right),
            Paragraph(f"Rs.{s['unit_price']:.2f}", style_cell_right),
            Paragraph(f"Rs.{s['purchase_price_at_time']:.2f}", style_cell_right),
            Paragraph(f"Rs.{s['total_revenue']:.2f}", style_cell_right),
            Paragraph(f"Rs.{s['margin_profit']:.2f}", style_cell_right_bold)
        ]
        sales_rows.append(row_cells)

    # Sales Totals row
    sales_rows.append([
        Paragraph("GRAND TOTALS", style_cell_bold),
        Paragraph(f"{total_qty:.1f} Units" if total_qty % 1 != 0 else f"{int(total_qty)} Units", style_cell_right_bold),
        Paragraph("", style_cell_normal),
        Paragraph("", style_cell_normal),
        Paragraph(f"Rs.{total_rev:.2f}", style_cell_right_bold),
        Paragraph(f"Rs.{total_gp:.2f}", style_cell_right_bold)
    ])

    # 7.5cm for desc, 1.7cm qty, 2.2cm rates, 2.2cm total rev, 2.2cm margin
    sales_table = Table(sales_rows, colWidths=[7.0*cm, 1.8*cm, 2.3*cm, 2.3*cm, 2.3*cm, 2.3*cm])
    
    sales_table_style = [
        ('BACKGROUND', (0,0), (-1,0), c_dark),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, c_border),
    ]
    
    # Alternating row colors
    for i in range(1, len(sales_rows) - 1):
        if i % 2 == 0:
            sales_table_style.append(('BACKGROUND', (0, i), (-1, i), c_light))
            
    # Bottom summary row highlights
    sales_table_style.extend([
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#F2EFE9")),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, c_brand),
    ])
    
    sales_table.setStyle(TableStyle(sales_table_style))
    story.append(sales_table)
    story.append(Spacer(1, 0.4*cm))

    # ----------------------------------------------------
    # SECTION 2: Daily Expenses Table
    # ----------------------------------------------------
    expense_data_list = []
    expense_data_list.append([
        Paragraph("Operating Expenditure Category", style_cell_header),
        Paragraph("Detailed Note / Description", style_cell_header),
        Paragraph("Amount (Rs.)", style_cell_header_right)
    ])
    
    for idx, e in enumerate(expenses):
        row_cells = [
            Paragraph(e['category'], style_cell_bold),
            Paragraph(e['note'] or "-", style_cell_normal),
            Paragraph(f"Rs.{e['amount']:.2f}", style_cell_right_bold)
        ]
        expense_data_list.append(row_cells)
        
    expense_data_list.append([
        Paragraph("TOTAL BUSINESS OPERATING EXPENDITURES", style_cell_bold),
        Paragraph("", style_cell_normal),
        Paragraph(f"Rs.{total_exp:.2f}", style_cell_right_bold)
    ])
    
    exp_table = Table(expense_data_list, colWidths=[6.0*cm, 8.5*cm, 3.5*cm])
    exp_table_style = [
        ('BACKGROUND', (0,0), (-1,0), c_dark),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, c_border),
    ]
    for i in range(1, len(expense_data_list) - 1):
        if i % 2 == 0:
            exp_table_style.append(('BACKGROUND', (0, i), (-1, i), c_light))
    exp_table_style.extend([
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FDF2F1")),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, c_accent_red),
    ])
    exp_table.setStyle(TableStyle(exp_table_style))
    
    # Pack operating costs log inside KeepTogether to prevent visual splits
    story.append(KeepTogether([
        Paragraph("2. DAILY OPERATING EXPENDITURES", style_h1),
        exp_table
    ]))
    story.append(Spacer(1, 0.4*cm))

    # ----------------------------------------------------
    # SECTION 3: Inventory Reconciliation Ledger Table
    # ----------------------------------------------------
    ledger_rows = [[
        Paragraph("Product Specification Name", style_cell_header),
        Paragraph("Opening", style_cell_header_right),
        Paragraph("Intake (+)", style_cell_header_right),
        Paragraph("Dispatch (-)", style_cell_header_right),
        Paragraph("Closing Stock", style_cell_header_right),
        Paragraph("Reconciliation Status", style_cell_header)
    ]]

    for idx, row in enumerate(ledger):
        desc = f"{row['product_name']} ({row['product_brand']}) - {row['product_thickness']}"
        op_text = f"{row['opening_stock']:.1f}" if row['opening_stock'] % 1 != 0 else f"{int(row['opening_stock'])}"
        pur_text = f"{row['purchases_qty']:.1f}" if row['purchases_qty'] % 1 != 0 else f"{int(row['purchases_qty'])}"
        sal_text = f"{row['sales_qty']:.1f}" if row['sales_qty'] % 1 != 0 else f"{int(row['sales_qty'])}"
        cls_text = f"{row['closing_stock']:.1f}" if row['closing_stock'] % 1 != 0 else f"{int(row['closing_stock'])}"
        
        status_text = "CONFIRMED & LOCKED" if row['is_confirmed'] == 1 else "RECONCILED / ADJUSTED" if row['manual_override'] == 1 else "UNCONFIRMED STOCK"
        if row['manual_override'] == 1 and row['override_note']:
             status_text += f"\n({row['override_note']})"
             
        row_cells = [
            Paragraph(desc, style_cell_normal),
            Paragraph(f"{op_text} {row['product_unit']}", style_cell_right),
            Paragraph(f"{pur_text} {row['product_unit']}" if row['purchases_qty'] > 0 else "-", style_cell_right),
            Paragraph(f"{sal_text} {row['product_unit']}" if row['sales_qty'] > 0 else "-", style_cell_right),
            Paragraph(f"{cls_text} {row['product_unit']}", style_cell_right_bold),
            Paragraph(status_text, style_cell_bold if row['is_confirmed'] == 1 else style_cell_normal)
        ]
        ledger_rows.append(row_cells)

    ledger_table = Table(ledger_rows, colWidths=[6.0*cm, 2.0*cm, 2.0*cm, 2.0*cm, 2.0*cm, 4.0*cm])
    ledger_table_style = [
        ('BACKGROUND', (0,0), (-1,0), c_dark),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, c_border),
    ]
    for i in range(1, len(ledger_rows)):
        row = ledger[i-1]
        if row['is_confirmed'] == 1:
            bg_col = colors.HexColor("#E8F8F5")  # light green
        elif row['manual_override'] == 1:
            bg_col = colors.HexColor("#EBF5FB")  # light blue
        elif i % 2 == 0:
            bg_col = c_light
        else:
            bg_col = colors.white
        ledger_table_style.append(('BACKGROUND', (0, i), (-1, i), bg_col))

    ledger_table.setStyle(TableStyle(ledger_table_style))

    story.append(KeepTogether([
        Paragraph("3. PHYSICAL STOCK RECONCILIATION SUMMARY", style_h1),
        ledger_table
    ]))

    # Footer signature line
    story.append(Spacer(1, 1.2*cm))
    sig_data = [
        [Paragraph("Report Generated Electronically", style_cell_normal), Paragraph("Authorized Store Manager Signature", style_cell_right)],
        [Paragraph("", style_cell_normal), Paragraph("_____________________________________", style_cell_right)]
    ]
    sig_table = Table(sig_data, colWidths=[9.0*cm, 9.0*cm])
    sig_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
    ]))
    
    story.append(KeepTogether([sig_table]))

    # Build Document
    doc.build(story)
    return True
